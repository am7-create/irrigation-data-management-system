#!/usr/bin/env python3
"""Check Excel rainfall values using openpyxl"""

import pandas as pd
import os
import glob
from openpyxl import load_workbook

# Use the correct data folder path
DATA_FOLDER = r"C:\Users\KIIT0001\Desktop\irrigation data management system\RAINFALL RIVER GAUGE 2025 WMD\RAINFALL RIVER GAUGE 2025 WMD"

# Find first Excel file
files = glob.glob(os.path.join(DATA_FOLDER, '**', '*.xlsx'), recursive=True)
files = [f for f in files if not os.path.basename(f).startswith('~$')]

if files:
    file = files[0]
    print(f'Checking file: {file}')
    
    # Try with openpyxl directly
    wb = load_workbook(file)
    ws = wb['Rainfall']
    
    print('\n--- Using openpyxl directly ---')
    print('Column F (rainfall) values (rows 5-15):')
    for row in range(5, 16):
        cell = ws[f'F{row}']
        print(f'  Row {row}: {repr(cell.value)} ({type(cell.value)})')
        
    wb.close()
    
    # Also try pandas with na_filter=False
    print('\n--- Pandas with na_filter=False ---')
    df = pd.read_excel(file, sheet_name='Rainfall', header=None, na_filter=False)
    print('Column 5 (rainfall) values (rows 4-10):')
    for i in range(4, 11):
        val = df.iloc[i, 5] if i < len(df) else 'N/A'
        print(f'  Row {i}: {repr(val)} ({type(val)})')
        
else:
    print('No Excel files found')
    print(f'Searched in: {DATA_FOLDER}')
    print(f'Folder exists: {os.path.exists(DATA_FOLDER)}')